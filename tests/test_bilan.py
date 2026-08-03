from openpyxl import Workbook
from visualisation_performances_sportives.analyse_sportive.bilan import construire_page_bilan

def test_construire_page_bilan():
    """Vérifie que l'onglet Bilan est bien créé avec son titre et ses formules."""
    wb = Workbook()
    
    # On simule 150 lignes de données pour tester la formule
    construire_page_bilan(wb, max_data_row=150, len_y=5, len_s=4, len_n=3)
    
    assert "Bilan" in wb.sheetnames, "L'onglet 'Bilan' n'a pas été créé."
    
    ws = wb["Bilan"]
    assert ws['A1'].value == "BILAN GLOBAL", "Le titre principal est incorrect."
    assert ws['B5'].value == "Total Activités:", "Le libellé du KPI est manquant."
    assert "COUNTA(Data!A2:A150)" in ws['C5'].value, "La formule Excel générée est incorrecte."