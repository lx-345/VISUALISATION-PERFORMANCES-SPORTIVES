from openpyxl import Workbook
from visualisation_performances_sportives.analyse_sportive.charge import construire_page_charge
import pytest

@pytest.mark.unit
def test_construire_page_charge():
    """Vérifie que l'onglet Charge de travail est bien créé avec son titre."""
    wb = Workbook()
    
    construire_page_charge(wb, max_data_row=100, len_y=5, len_t=12, len_d=5)
    
    assert "Charge de travail" in wb.sheetnames, "L'onglet 'Charge de travail' n'a pas été créé."
    
    ws = wb["Charge de travail"]
    assert ws['A1'].value == "SUIVI DE LA CHARGE DE TRAVAIL", "Le titre principal est incorrect."