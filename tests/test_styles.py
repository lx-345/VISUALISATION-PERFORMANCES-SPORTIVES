from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border
from visualisation_performances_sportives.analyse_sportive.styles import get_base_styles, appliquer_theme_global

def test_get_base_styles():
    """Vérifie que les 5 éléments de style sont bien retournés et du bon type."""
    styles = get_base_styles()
    
    assert len(styles) == 5, "La fonction doit retourner exactement 5 éléments de style."
    assert isinstance(styles[0], PatternFill), "Le premier élément doit être un PatternFill."
    assert isinstance(styles[1], Font), "Le deuxième élément doit être une Font."
    assert isinstance(styles[2], Font), "Le troisième élément doit être une Font."
    assert isinstance(styles[3], Alignment), "Le quatrième élément doit être un Alignment."
    assert isinstance(styles[4], Border), "Le cinquième élément doit être une Border."

def test_appliquer_theme_global():
    """Vérifie que le fond de la feuille est bien modifié."""
    wb = Workbook()
    ws = wb.active
    
    appliquer_theme_global(ws)
    
    # On vérifie la couleur de fond de la cellule A1
    cellule_a1 = ws['A1']
    assert cellule_a1.fill.start_color.index == "001E1E1E", "La couleur de fond devrait être 1E1E1E."