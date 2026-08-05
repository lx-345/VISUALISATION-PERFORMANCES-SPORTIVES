import pandas as pd
import pytest
from openpyxl import Workbook

from visualisation_performances_sportives.analyse_sportive.backend import (
    construire_onglets_back,
    preparer_donnees_agregees,
)


@pytest.mark.unit
def test_preparer_donnees_agregees():
    """Vérifie que les calculs de regroupement par année fonctionnent."""
    df_test = pd.DataFrame(
        {
            "annee": [2023, 2023, 2024],
            "trimestre": ["T1", "T2", "T1"],
            "distance_km": [10.5, 15.0, 20.0],
        }
    )

    df_yearly, df_trim = preparer_donnees_agregees(df_test)

    valeur_2023 = df_yearly.loc[df_yearly["annee"] == 2023, "distance_km"].values[0]
    assert valeur_2023 == 25.5, (
        f"Erreur d'agrégation : attendu 25.5, reçu {valeur_2023}"
    )


def test_construire_onglets_back():
    """Vérifie que le backend crée bien les onglets cachés dans Excel."""
    wb = Workbook()
    df_test = pd.DataFrame(
        {"annee": [2023], "trimestre": ["T1"], "distance_km": [10.5]}
    )

    construire_onglets_back(wb, df_test)

    assert "Data" in wb.sheetnames, "L'onglet 'Data' n'a pas été créé."
    assert "Backend" in wb.sheetnames, "L'onglet 'Backend' n'a pas été créé."
    assert wb["Data"].sheet_state == "hidden", "L'onglet 'Data' devrait être caché."
