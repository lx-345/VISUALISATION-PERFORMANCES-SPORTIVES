import pytest
from openpyxl import Workbook

from visualisation_performances_sportives.analyse_sportive.analyse import (
    construire_page_analyse,
)


@pytest.mark.unit
def test_construire_page_analyse():
    """Vérifie que l'onglet Analyse est bien créé, titré et contient un graphique."""
    wb = Workbook()

    # On simule la présence de l'onglet Backend créé à l'étape précédente
    wb.create_sheet(title="Backend")

    # Paramètres factices pour simuler l'exécution
    construire_page_analyse(
        wb, max_data_row=100, len_y=5, len_t=12, len_d=5, max_pivot=10
    )

    # Vérifications
    assert "Analyse" in wb.sheetnames, "L'onglet 'Analyse' n'a pas été créé."

    ws = wb["Analyse"]
    assert ws["A1"].value == "ANALYSE DES PERFORMANCES SPORTIVES", (
        "Le titre principal est incorrect."
    )
    assert len(ws._charts) > 0, "Le graphique n'a pas été ajouté à la feuille."
