from openpyxl.chart import BarChart, Reference

from .styles import appliquer_theme_global, get_base_styles


def construire_page_analyse(wb, max_data_row, len_y, len_t, len_d, max_pivot):
    """
    Génère l'onglet principal 'Analyse' avec ses graphiques de performance.
    """
    # Création de l'onglet en première position (index 0)
    ws = wb.create_sheet(title="Analyse", index=0)
    appliquer_theme_global(ws)

    # Récupération de vos styles
    couleur_fond, police_titre, _, align_center, bordure = get_base_styles()

    # Mise en forme de l'en-tête
    ws["A1"] = "ANALYSE DES PERFORMANCES SPORTIVES"
    ws["A1"].font = police_titre
    ws["A1"].fill = couleur_fond
    ws["A1"].alignment = align_center
    ws["A1"].border = bordure
    ws.merge_cells("A1:J2")

    # Création d'un graphique (Distance annuelle)
    chart = BarChart()
    chart.title = "Distance Parcourue par Année"
    chart.style = 10

    # Sécurité : on s'assure que le Backend existe avant de lier les données
    if "Backend" in wb.sheetnames:
        data = Reference(wb["Backend"], min_col=2, min_row=1, max_row=len_y, max_col=2)
        cats = Reference(wb["Backend"], min_col=1, min_row=2, max_row=len_y)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

    # Placement du graphique
    ws.add_chart(chart, "A5")
