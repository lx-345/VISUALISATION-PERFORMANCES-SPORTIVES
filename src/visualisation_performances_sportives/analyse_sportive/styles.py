from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def get_base_styles():
    """
    Définit et retourne les styles standards (polices, couleurs, bordures)
    utilisés dans tout le tableau de bord Excel.
    """
    couleur_fond = PatternFill(
        start_color="333333", end_color="333333", fill_type="solid"
    )
    police_titre = Font(color="FFFFFF", bold=True, size=14)
    police_texte = Font(color="FFFFFF", size=11)
    alignement_centre = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    bordure_fine = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF"),
    )

    return couleur_fond, police_titre, police_texte, alignement_centre, bordure_fine


def appliquer_theme_global(ws):
    """
    Applique le fond sombre (mode dark) sur l'ensemble de la feuille de calcul.
    """
    fond_sombre = PatternFill(
        start_color="1E1E1E", end_color="1E1E1E", fill_type="solid"
    )
    # Pour le test et l'optimisation, on limite l'application à une zone raisonnable
    for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=20):
        for cell in row:
            cell.fill = fond_sombre
